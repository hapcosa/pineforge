"""
Generador de Excel profesional con formato condicional por estado.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from .config import EXCEL_COLUMNS, STATE_COLORS
except ImportError:  # Permite ejecutar python dvu_bot/main.py
    from config import EXCEL_COLUMNS, STATE_COLORS

logger = logging.getLogger(__name__)


HEADER_FILL = PatternFill(start_color="FF1E293B", end_color="FF1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
CELL_FONT = Font(name="Calibri", size=10)
THIN = Side(border_style="thin", color="FFCCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _auto_width(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)


def write_excel(records: List[Dict], output_path: Path) -> None:
    """Genera Excel con header estilizado, colores por estado y ancho automatico."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    # Header
    for c_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    # Filas
    for r_idx, rec in enumerate(records, start=2):
        estado = rec.get("Estado", "")
        fill_color = STATE_COLORS.get(estado)
        row_fill = (
            PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if fill_color
            else None
        )

        for c_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
            val = rec.get(col_name, "")
            # Truncar OCR demasiado largo (Excel admite 32767 pero legibilidad)
            if col_name == "Texto OCR" and isinstance(val, str) and len(val) > 2000:
                val = val[:2000] + "…"
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if row_fill is not None:
                cell.fill = row_fill

    # Autofiltro + ancho
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{max(1, len(records) + 1)}"
    _auto_width(ws, len(EXCEL_COLUMNS))

    wb.save(output_path)
    logger.info(f"Excel generado: {output_path} ({len(records)} filas)")
