"""
Reconciliacion de comprobantes DVU contra cartola bancaria.
"""

from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from .config import CARTOLA_XLSX, OUTPUT_XLSX, RECONCILIACION_XLSX, STATE_COLORS
    from .excel import _auto_width
    from .extractor import _strip_accents, extract_facturas
except ImportError:  # Permite importar desde scripts locales dentro de dvu_bot/
    from config import CARTOLA_XLSX, OUTPUT_XLSX, RECONCILIACION_XLSX, STATE_COLORS
    from excel import _auto_width
    from extractor import _strip_accents, extract_facturas


logger = logging.getLogger(__name__)

SHEET_ORDER = [
    "RESUMEN",
    "PAGOS_OK",
    "SIN_COMPROBANTE",
    "TERCERO_PROBABLE",
    "REVISION_MANUAL",
    "COMPROBANTES_HUERFANOS",
]

SHEET_COLORS = {
    "PAGOS_OK": STATE_COLORS["LISTO PARA INGRESAR"],
    "SIN_COMPROBANTE": STATE_COLORS["FALTA DATO"],
    "TERCERO_PROBABLE": STATE_COLORS["ABONO PARCIAL"],
    "REVISION_MANUAL": STATE_COLORS["DUPLICADO POSIBLE"],
    "COMPROBANTES_HUERFANOS": STATE_COLORS["REVISAR OCR"],
}

HEADER_FILL = PatternFill(start_color="FF1E293B", end_color="FF1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
CELL_FONT = Font(name="Calibri", size=10)
THIN = Side(border_style="thin", color="FFCCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LEGAL_NAME_STOPWORDS = {
    "SPA",
    "S",
    "P",
    "A",
    "LIMITADA",
    "LTDA",
    "SOC",
    "SOCIEDAD",
    "SR",
    "SRA",
    "SENOR",
    "SENORA",
    "DE",
    "DEL",
    "LA",
    "EL",
    "LOS",
    "LAS",
    "Y",
    "E",
    "POR",
    "PARA",
    "PAGO",
    "PAGOS",
    "FACT",
    "FACTURA",
    "FACTURAS",
    "CLIENTE",
    "DON",
    "DONA",
    "REVISAR",
}


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except TypeError:
        pass
    return str(value).strip() == ""


def _as_text(value: Any) -> str:
    if _is_blank(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.date().strftime("%d/%m/%Y")
    if isinstance(value, datetime):
        return value.date().strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_monto(value: Any) -> Optional[int]:
    """Normaliza montos tipo '$ 1.000.000' a int."""
    if _is_blank(value):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value)) if not pd.isna(value) else None

    text = str(value).strip()
    text = re.sub(r",\d{1,2}$", "", text)
    digits = re.sub(r"\D", "", text)
    return int(digits) if digits else None


def norm_nombre(value: Any) -> str:
    """Normaliza nombres para comparacion tolerante."""
    if _is_blank(value):
        return ""
    text = _strip_accents(str(value)).upper()
    text = re.sub(r"\bS\s*P\s*A\b", " SPA ", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    tokens = [
        token for token in text.split()
        if token not in LEGAL_NAME_STOPWORDS and len(token) > 1
    ]
    return " ".join(tokens)


def norm_fecha(value: Any) -> Optional[date]:
    """Normaliza fechas dd/mm/yy, dd/mm/yyyy o valores Excel a date."""
    if _is_blank(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y", "%d.%m.%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _norm_header(value: Any) -> str:
    if _is_blank(value):
        return ""
    text = _strip_accents(str(value)).upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return " ".join(text.split())


def _dedupe_columns(values: Sequence[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    columns: List[str] = []
    for idx, value in enumerate(values, start=1):
        base = _as_text(value) or f"Columna {idx}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        columns.append(base if count == 0 else f"{base}_{count + 1}")
    return columns


def _detect_header_row(raw: pd.DataFrame) -> int:
    for idx, row in raw.iterrows():
        normalized = " ".join(_norm_header(value) for value in row.tolist())
        if "FECHA" in normalized and "MONTO" in normalized:
            return int(idx)
    return 0


def _read_cartola(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, header=None, dtype=object)
    header_idx = _detect_header_row(raw)
    columns = _dedupe_columns(raw.iloc[header_idx].tolist())
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = columns
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def _read_comprobantes(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=object)
    return df.dropna(how="all").reset_index(drop=True)


def _find_col(columns: Iterable[str], groups: Sequence[Sequence[str]]) -> Optional[str]:
    normalized = [(col, _norm_header(col)) for col in columns]
    for group in groups:
        wanted = [_strip_accents(part).upper() for part in group]
        for col, norm in normalized:
            if all(part in norm for part in wanted):
                return col
    return None


def _prepare_cartola(df: pd.DataFrame) -> List[Dict[str, Any]]:
    fecha_col = _find_col(df.columns, [("FECHA",)])
    monto_col = _find_col(df.columns, [("MONTO",)])
    nombre_col = _find_col(
        df.columns,
        [("NOMBRE", "ORIGEN"), ("NOMBRE", "DESTINO"), ("NOMBRE",)],
    )
    if not fecha_col or not monto_col:
        raise ValueError("No pude detectar columnas Fecha y Monto en la cartola.")

    rows: List[Dict[str, Any]] = []
    for idx, row in df.iterrows():
        item = row.to_dict()
        item["_id"] = int(idx)
        item["_monto"] = norm_monto(row.get(monto_col))
        item["_nombre"] = norm_nombre(row.get(nombre_col)) if nombre_col else ""
        item["_fecha"] = norm_fecha(row.get(fecha_col))
        rows.append(item)
    return rows


def _prepare_comprobantes(df: pd.DataFrame) -> List[Dict[str, Any]]:
    monto_col = _find_col(df.columns, [("MONTO", "TRANSFERIDO"), ("MONTO",)])
    nombre_col = _find_col(df.columns, [("CLIENTE", "DETECTADO"), ("CLIENTE",), ("NOMBRE",)])
    fecha_transferencia_col = _find_col(df.columns, [("FECHA", "TRANSFERENCIA")])
    fecha_mensaje_col = _find_col(df.columns, [("FECHA", "MENSAJE")])
    fecha_col = fecha_transferencia_col or fecha_mensaje_col or _find_col(df.columns, [("FECHA",)])
    if not monto_col:
        raise ValueError("No pude detectar columna de monto en comprobantes.")

    rows: List[Dict[str, Any]] = []
    for idx, row in df.iterrows():
        item = row.to_dict()
        item["_id"] = int(idx)
        item["_monto"] = norm_monto(row.get(monto_col))
        item["_nombre"] = norm_nombre(row.get(nombre_col)) if nombre_col else ""
        fecha = norm_fecha(row.get(fecha_transferencia_col)) if fecha_transferencia_col else None
        if fecha is None and fecha_mensaje_col:
            fecha = norm_fecha(row.get(fecha_mensaje_col))
        if fecha is None and fecha_col:
            fecha = norm_fecha(row.get(fecha_col))
        item["_fecha"] = fecha
        rows.append(item)
    return rows


def _nombre_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if len(a) >= 4 and a in b:
        return True
    if len(b) >= 4 and b in a:
        return True
    return SequenceMatcher(None, a, b).ratio() >= 0.70


def _nombre_score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a in b or b in a:
        return 1.0
    return SequenceMatcher(None, a, b).ratio()


def _fecha_cerca(a: Optional[date], b: Optional[date], days: int = 2) -> bool:
    if not a or not b:
        return False
    return abs((a - b).days) <= days


def _amount_sum(rows: Iterable[Dict[str, Any]]) -> int:
    return sum(row.get("_monto") or 0 for row in rows)


def _prefix_row(row: Dict[str, Any], prefix: str) -> Dict[str, Any]:
    return {
        f"{prefix}{key}": "" if key.startswith("_") else value
        for key, value in row.items()
        if not key.startswith("_")
    }


def _combine_rows(cartola: Dict[str, Any], comp: Dict[str, Any], extra: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    out.update(extra)
    out["Monto Normalizado"] = cartola.get("_monto") or comp.get("_monto") or ""
    out["Nombre Cartola Normalizado"] = cartola.get("_nombre", "")
    out["Nombre Comprobante Normalizado"] = comp.get("_nombre", "")
    out["Factura(s) OCR"] = _facturas_ocr(comp)
    out.update(_prefix_row(cartola, "Cartola - "))
    out.update(_prefix_row(comp, "Comprobante - "))
    return out


def _cartola_only(row: Dict[str, Any], extra: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(extra)
    out["Monto Normalizado"] = row.get("_monto") or ""
    out["Nombre Cartola Normalizado"] = row.get("_nombre", "")
    out.update(_prefix_row(row, "Cartola - "))
    return out


def _comp_only(row: Dict[str, Any], extra: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(extra)
    out["Monto Normalizado"] = row.get("_monto") or ""
    out["Nombre Comprobante Normalizado"] = row.get("_nombre", "")
    out["Factura(s) OCR"] = _facturas_ocr(row)
    out.update(_prefix_row(row, "Comprobante - "))
    return out


def _facturas_ocr(row: Dict[str, Any]) -> str:
    if not _is_blank(row.get("Factura(s)")):
        return ""
    facturas = extract_facturas(_as_text(row.get("Texto OCR")))
    return ", ".join(facturas)


def _describe_cartola(row: Dict[str, Any]) -> str:
    return (
        f"cartola#{row['_id'] + 1} | fecha={_as_text(row.get('_fecha'))} | "
        f"monto={row.get('_monto') or ''} | nombre={row.get('_nombre')}"
    )


def _describe_comp(row: Dict[str, Any]) -> str:
    archivo = row.get("Archivo Imagen", "")
    facturas = row.get("Factura(s)", "")
    op = row.get("N° Operación", "")
    return (
        f"comp#{row['_id'] + 1} | fecha={_as_text(row.get('_fecha'))} | "
        f"monto={row.get('_monto') or ''} | nombre={row.get('_nombre')} | "
        f"archivo={_as_text(archivo)} | facturas={_as_text(facturas)} | op={_as_text(op)}"
    )


def _best_match(cartola: Dict[str, Any], candidates: Sequence[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    matches = [
        comp for comp in candidates
        if _nombre_match(cartola.get("_nombre", ""), comp.get("_nombre", ""))
        and _fecha_cerca(cartola.get("_fecha"), comp.get("_fecha"))
    ]
    if not matches:
        return None
    return max(
        matches,
        key=lambda comp: (
            _nombre_score(cartola.get("_nombre", ""), comp.get("_nombre", "")),
            -abs((cartola["_fecha"] - comp["_fecha"]).days),
        ),
    )


def _rows_to_df(rows: List[Dict[str, Any]], columns: Sequence[str]) -> pd.DataFrame:
    if rows:
        return pd.DataFrame(rows)
    return pd.DataFrame(columns=list(columns))


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        max_col = max(ws.max_column, 1)
        max_row = max(ws.max_row, 1)

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        sheet_fill = None
        if ws.title in SHEET_COLORS:
            color = SHEET_COLORS[ws.title]
            sheet_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.font = CELL_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = BORDER
                if sheet_fill is not None:
                    cell.fill = sheet_fill

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        _auto_width(ws, max_col)

    wb.save(path)


def reconciliar(
    comprobantes_path: Path = OUTPUT_XLSX,
    cartola_path: Path = CARTOLA_XLSX,
    output_path: Path = RECONCILIACION_XLSX,
) -> Path:
    """Cruza comprobantes contra cartola y genera reconciliacion_dvu.xlsx."""
    comprobantes_path = Path(comprobantes_path)
    cartola_path = Path(cartola_path)
    output_path = Path(output_path)

    if not comprobantes_path.exists():
        raise FileNotFoundError(f"No existe el archivo de comprobantes: {comprobantes_path}")
    if not cartola_path.exists():
        raise FileNotFoundError(f"No existe la cartola: {cartola_path}")

    cartola = _prepare_cartola(_read_cartola(cartola_path))
    comprobantes = _prepare_comprobantes(_read_comprobantes(comprobantes_path))

    cartola_by_amount: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    comp_by_amount: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for row in cartola:
        if row.get("_monto") is not None:
            cartola_by_amount[row["_monto"]].append(row)
    for row in comprobantes:
        if row.get("_monto") is not None:
            comp_by_amount[row["_monto"]].append(row)

    pagos_ok: List[Dict[str, Any]] = []
    sin_comprobante: List[Dict[str, Any]] = []
    tercero_probable: List[Dict[str, Any]] = []
    revision_manual: List[Dict[str, Any]] = []

    assigned_cartola: set[int] = set()
    assigned_comp: set[int] = set()
    manual_cartola: set[int] = set()
    manual_comp: set[int] = set()
    manual_amounts: set[int] = set()

    def add_manual(amount: int, motivo: str) -> None:
        if amount in manual_amounts:
            return
        manual_amounts.add(amount)
        cart_rows = cartola_by_amount.get(amount, [])
        comp_rows = comp_by_amount.get(amount, [])
        manual_cartola.update(row["_id"] for row in cart_rows)
        manual_comp.update(row["_id"] for row in comp_rows)
        revision_manual.append({
            "Monto": amount,
            "Cartola Candidatos": len(cart_rows),
            "Comprobantes Candidatos": len(comp_rows),
            "Candidatos Cartola": "\n".join(_describe_cartola(row) for row in cart_rows),
            "Candidatos Comprobantes": "\n".join(_describe_comp(row) for row in comp_rows),
            "Motivo": motivo,
            "Accion": "",
        })

    for cart_row in cartola:
        amount = cart_row.get("_monto")
        if amount is None:
            sin_comprobante.append(_cartola_only(cart_row, {"Motivo": "Monto no detectado en cartola"}))
            assigned_cartola.add(cart_row["_id"])
            continue

        candidates = comp_by_amount.get(amount, [])
        if not candidates:
            sin_comprobante.append(_cartola_only(cart_row, {"Motivo": "Sin comprobante con monto exacto"}))
            assigned_cartola.add(cart_row["_id"])
            continue

        if len(candidates) == 1:
            comp = candidates[0]
            if comp["_id"] in assigned_comp:
                add_manual(amount, "Monto ya asociado a otro registro; revisar duplicidad")
                continue

            name_ok = _nombre_match(cart_row.get("_nombre", ""), comp.get("_nombre", ""))
            date_ok = _fecha_cerca(cart_row.get("_fecha"), comp.get("_fecha"))
            if name_ok and date_ok:
                pagos_ok.append(_combine_rows(cart_row, comp, {"Match": "Nombre y fecha +/-2 dias"}))
                assigned_cartola.add(cart_row["_id"])
                assigned_comp.add(comp["_id"])
            elif (
                date_ok
                and (not cart_row.get("_nombre") or not comp.get("_nombre"))
                and len(cartola_by_amount.get(amount, [])) == 1
                and len(comp_by_amount.get(amount, [])) == 1
            ):
                pagos_ok.append(_combine_rows(cart_row, comp, {
                    "Match": "Monto unico y fecha +/-2 dias; nombre ausente",
                }))
                assigned_cartola.add(cart_row["_id"])
                assigned_comp.add(comp["_id"])
            elif (
                not name_ok
                and len(cartola_by_amount.get(amount, [])) == 1
                and len(comp_by_amount.get(amount, [])) == 1
            ):
                mensaje = (
                    f"tercero pago: {_as_text(comp.get('Cliente Detectado')) or comp.get('_nombre', '')} "
                    f"por {_as_text(cart_row.get('Nombre Destino/Origen')) or cart_row.get('_nombre', '')}"
                )
                tercero_probable.append(_combine_rows(cart_row, comp, {
                    "Mensaje": mensaje,
                    "Motivo": "Monto unico en cartola y comprobantes; nombre distinto",
                }))
                assigned_cartola.add(cart_row["_id"])
                assigned_comp.add(comp["_id"])
            else:
                add_manual(amount, "Un candidato por monto, pero nombre/fecha no confirma")
            continue

        best = _best_match(cart_row, [comp for comp in candidates if comp["_id"] not in assigned_comp])
        if best is not None:
            pagos_ok.append(_combine_rows(cart_row, best, {"Match": "Monto repetido, nombre y fecha confirman"}))
            assigned_cartola.add(cart_row["_id"])
            assigned_comp.add(best["_id"])
            add_manual(amount, "Monto repetido; revisar candidatos no asociados automaticamente")
        else:
            add_manual(amount, "Monto repetido sin confirmacion por nombre y fecha")

    comprobantes_huerfanos = [
        _comp_only(row, {"Motivo": "Comprobante sin match en cartola"})
        for row in comprobantes
        if row["_id"] not in assigned_comp and row["_id"] not in manual_comp
    ]

    manual_cart_rows = [row for row in cartola if row["_id"] in manual_cartola]
    manual_comp_rows = [row for row in comprobantes if row["_id"] in manual_comp]
    orphan_comp_rows = [
        row for row in comprobantes
        if row["_id"] not in assigned_comp and row["_id"] not in manual_comp
    ]

    resumen = [
        {
            "Categoria": "PAGOS_OK",
            "Registros": len(pagos_ok),
            "Total Cartola $": sum(row.get("Monto Normalizado") or 0 for row in pagos_ok),
            "Total Comprobantes $": sum(row.get("Monto Normalizado") or 0 for row in pagos_ok),
            "Notas": "Monto, nombre y fecha conciliados.",
        },
        {
            "Categoria": "SIN_COMPROBANTE",
            "Registros": len(sin_comprobante),
            "Total Cartola $": sum(row.get("Monto Normalizado") or 0 for row in sin_comprobante),
            "Total Comprobantes $": 0,
            "Notas": "Depositos en cartola sin comprobante exacto por monto.",
        },
        {
            "Categoria": "TERCERO_PROBABLE",
            "Registros": len(tercero_probable),
            "Total Cartola $": sum(row.get("Monto Normalizado") or 0 for row in tercero_probable),
            "Total Comprobantes $": sum(row.get("Monto Normalizado") or 0 for row in tercero_probable),
            "Notas": "Monto unico con nombre distinto.",
        },
        {
            "Categoria": "REVISION_MANUAL",
            "Registros": len(revision_manual),
            "Total Cartola $": _amount_sum(manual_cart_rows),
            "Total Comprobantes $": _amount_sum(manual_comp_rows),
            "Notas": "Montos repetidos o fecha/nombre sin confirmacion.",
        },
        {
            "Categoria": "COMPROBANTES_HUERFANOS",
            "Registros": len(comprobantes_huerfanos),
            "Total Cartola $": 0,
            "Total Comprobantes $": _amount_sum(orphan_comp_rows),
            "Notas": "Comprobantes enviados que no aparecen conciliados.",
        },
    ]

    sheets = {
        "RESUMEN": pd.DataFrame(resumen),
        "PAGOS_OK": _rows_to_df(pagos_ok, ["Match", "Monto Normalizado"]),
        "SIN_COMPROBANTE": _rows_to_df(sin_comprobante, ["Motivo", "Monto Normalizado"]),
        "TERCERO_PROBABLE": _rows_to_df(tercero_probable, ["Mensaje", "Motivo", "Monto Normalizado"]),
        "REVISION_MANUAL": _rows_to_df(
            revision_manual,
            ["Monto", "Cartola Candidatos", "Comprobantes Candidatos", "Candidatos Cartola", "Candidatos Comprobantes", "Motivo", "Accion"],
        ),
        "COMPROBANTES_HUERFANOS": _rows_to_df(
            comprobantes_huerfanos,
            ["Motivo", "Monto Normalizado", "Nombre Comprobante Normalizado"],
        ),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet in SHEET_ORDER:
            sheets[sheet].to_excel(writer, sheet_name=sheet, index=False)

    _style_workbook(output_path)
    logger.info(f"Reconciliacion generada: {output_path}")
    return output_path
